#!/usr/bin/env python3
"""
AI Flow Studio — Content Prompt Generator v2.0
Python 3.12 | Windows 10
Usage:
    python generator.py       → 3 дня (по умолчанию)
    python generator.py 7     → 7 дней
Requires: pip install openpyxl
"""
import sys, random
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

DNA_FILE      = "CONTENT_DNA.xlsx"
OUTPUT_DIR    = Path("output")
DEFAULT_DAYS  = 3
ATOMS_PER_DAY = 4
COOLDOWN_DAYS = 14
MAX_MODE_DAY  = 2

PALETTE = [
    ["1E3A5F","DBEAFE","EFF6FF"],
    ["14532D","BBF7D0","F0FDF4"],
    ["78350F","FDE68A","FFFBEB"],
]
TASK_CLR = {
    "Текст озвучки":   "F3E8FF",
    "Стиль озвучки":   "EDE9FE",
    "Промт видео":     "FEE2E2",
    "Промт картинки":  "DCFCE7",
    "Подпись":         "DBEAFE",
}
WHITE = "FFFFFF"

COLS = [
    ("ДАТА",11),("АТОМ",7),("ШАГ",22),("ПЛАТФОРМА",14),
    ("ЗАДАЧА",22),("ПРОМТ — СКОПИРОВАТЬ И ВСТАВИТЬ В ИНСТРУМЕНТ",85),
    ("ИНСТРУМЕНТ",20),("✓",5),
]

# ── helpers ───────────────────────────────────────────────────────
def mkfill(c): return PatternFill("solid",fgColor=c)
def mkfont(bold=False,size=10,color="1E293B"): return Font(bold=bold,size=size,color=color)
def mkalign(h="left"): return Alignment(horizontal=h,vertical="top",wrap_text=True)
def mkcenter(): return Alignment(horizontal="center",vertical="center",wrap_text=True)

def load_sheet(wb, name):
    if name not in wb.sheetnames:
        print(f"  ВНИМАНИЕ: лист '{name}' не найден"); return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows)<2: return []
    headers=[str(h or "").strip() for h in rows[0]]
    return [{headers[i]:row[i] for i in range(len(headers))}
            for row in rows[1:] if any(v is not None for v in row)]

# ── anti-repeat ───────────────────────────────────────────────────
def parse_log(rows):
    out=[]
    for r in rows:
        raw=r.get("Дата")
        if not raw: continue
        try:
            d=raw if isinstance(raw,datetime) else datetime.strptime(str(raw).strip(),"%Y-%m-%d")
            out.append({"date":d,"sit_id":str(r.get("ИД_Ситуации","")),"mode":str(r.get("Режим","")),"industry":str(r.get("Индустрия",""))})
        except: pass
    return out

def used_ids(history, today):
    cut=today-timedelta(days=COOLDOWN_DAYS)
    return {h["sit_id"] for h in history if h["date"]>=cut}

def used_inds(history, today, days=2):
    cut=today-timedelta(days=days)
    return {h["industry"] for h in history if h["date"]>=cut}

def select_atoms(situations, modes_idx, history, today):
    blocked=used_ids(history,today); blocked_ind=used_inds(history,today)
    pool=[s for s in situations if str(s.get("ИД","")) not in blocked]
    if len(pool)<ATOMS_PER_DAY: pool=list(situations)
    scored=[]
    for s in pool:
        mode_id=str(s.get("Режим","Humor"))
        w=int((modes_idx.get(mode_id) or {}).get("Вес",10) or 10)
        score=w+int(s.get("Визуал",3) or 3)+int(s.get("Юмор",3) or 3)
        if str(s.get("Индустрия","")) not in blocked_ind: score+=3
        score+=random.uniform(0,3)
        scored.append((score,s))
    scored.sort(key=lambda x:-x[0])
    selected=[]; used_modes=[]
    for _,s in scored:
        if len(selected)>=ATOMS_PER_DAY: break
        m=str(s.get("Режим",""))
        if used_modes.count(m)>=MAX_MODE_DAY: continue
        selected.append(s); used_modes.append(m)
    if len(selected)<ATOMS_PER_DAY:
        rest=[s for _,s in scored if s not in selected]
        selected+=rest[:ATOMS_PER_DAY-len(selected)]
    return selected[:ATOMS_PER_DAY]

# ── tool resolver ─────────────────────────────────────────────────
def tool_map(tools):
    idx={}
    for t in sorted(tools,key=lambda x:int(x.get("Приоритет",99) or 99)):
        pt=str(t.get("Тип","")).lower()
        if pt not in idx: idx[pt]=t.get("Инструмент","—")
    return idx

def pick_image_tool(tools, has_thai):
    ranked=sorted([t for t in tools if str(t.get("Тип","")).lower()=="image"],
                  key=lambda x:int(x.get("Приоритет",99) or 99))
    if has_thai:
        for t in ranked:
            if "ideogram" in str(t.get("Инструмент","")).lower():
                return t["Инструмент"]+" ← тайский текст"
    return ranked[0]["Инструмент"] if ranked else "Ideogram v3"

# ══════════════════════════════════════════════════════════════════
#  ПРОМТЫ — ОСНОВНОЕ ЗНАЧЕНИЕ СИСТЕМЫ
# ══════════════════════════════════════════════════════════════════

BRAND_BLOCK = """AI Flow Studio — компания, которая помогает превращать ручные процессы в системы.
Контент показывает скрытые потери и абсурдность ручной работы — без корпоративного пафоса.
Никаких обещаний «AI изменит мир». Только конкретные ситуации и реальные результаты."""

MODE_INSTRUCTIONS = {
    "Humor": "Смеёмся над АБСУРДНОСТЬЮ СИТУАЦИИ — не над человеком. Логика: проблема → нарастание до абсурда → финальный инсайт.",
    "Serious": "Серьёзный анализ реальных потерь. Показываем факты, числа, последствия. Без сенсаций.",
    "Demo": "Показываем результат автоматизации в действии. До → процесс → после. Конкретно и понятно.",
    "Educational": "Объясняем просто — как умный друг, не как лектор. Шаг за шагом, с примерами.",
    "Before_After": "Контраст: как было (хаос, потери) → как стало (система, контроль). Эмоция: облегчение.",
    "Myth_Busting": "Берём распространённое заблуждение → разбиваем фактами → даём новую картину мира.",
    "Story": "Нарратив через узнаваемого персонажа. Проблема нарастает → точка перелома → выход.",
    "POV": "От первого лица: читатель = герой. Интимно, честно, без дистанции.",
}

EMOTION_ARCS = {
    "Humor":       "усмешка узнавания → нарастание абсурда → инсайт с улыбкой",
    "Serious":     "тревога от осознания → понимание масштаба → ясный вывод",
    "Demo":        "интерес → удивление от простоты → желание так же",
    "Educational": "вопрос → ступени понимания → уверенность",
    "Before_After":"усталость/смирение → контраст → облегчение",
    "Myth_Busting":"сомнение → разоблачение → новое понимание",
    "Story":       "сочувствие → нарастание → катарсис",
    "POV":         "раздражение/перегрузка → момент осознания",
}

VO_STYLES = {
    "Humor":       "Тёплый рассказчик с иронией. Как будто делишься смешной-но-правдивой историей с другом.",
    "Serious":     "Чёткий, уверенный, взвешенный. Авторитетный, но не холодный.",
    "Demo":        "Энергичный и понятный. Как будто показываешь другу что-то крутое.",
    "Educational": "Терпеливый эксперт. Объясняет в правильном темпе, не снисходительно.",
    "Before_After":"СТАРТ: усталый/смирившийся. КОНЕЦ: облегчённый/заряженный. Контраст — это вся история.",
    "Myth_Busting":"Слегка провокационный старт → уверенное развенчание → спокойный вывод.",
    "Story":       "Рассказчик документального фильма. Тёплый, с паузами, с эмоциональными точками.",
    "POV":         "Внутренний монолог от первого лица. Интимно, честно.",
}

def p_voiceover_script(sit, mode):
    """Промт для написания текста озвучки через Claude/GPT."""
    sit_text  = sit.get("Ситуация","")
    hook      = sit.get("Хук","")
    mode_id   = str(mode.get("ИД","Humor"))
    duration  = int(mode.get("Длит_сек",45) or 45)
    arc       = EMOTION_ARCS.get(mode_id,"нарастание → вывод")
    mode_inst = MODE_INSTRUCTIONS.get(mode_id,"")
    words     = duration * 2   # тайский: ~2 слова/сек

    hook_line = f"\nХУК (отправная точка, можно улучшить):\n{hook}" if hook else ""

    return f"""Напиши текст озвучки для видео AI Flow Studio.

БРЕНД:
{BRAND_BLOCK}

СИТУАЦИЯ:
{sit_text}
{hook_line}

РЕЖИМ: {mode.get("Режим","Юмор")}
{mode_inst}

ПАРАМЕТРЫ:
- Язык: тайский разговорный (не официальный, не литературный)
- Длительность: {duration} секунд / примерно {words} слов
- Эмоциональная дуга: {arc}

СТРУКТУРА С МАРКЕРАМИ ПАУЗЫ:
[0-3 сек]   ХУКА — первые слова обязаны остановить скролл
[PAUSE]
[4-{duration-10} сек]   РАЗВИТИЕ — углубляем ситуацию, нарастание
[BEAT]      (1 секунда тишины перед финалом)
[{duration-9}-{duration} сек]  ФИНАЛ — инсайт или открытый вопрос, не прямая реклама

НЕЛЬЗЯ:
- Корпоративный язык и AI-хайп
- «AI изменит ваш бизнес» / «инновационное решение»
- Унижение людей (смеёмся над ситуацией, не над человеком)
- Прямая продажа в финале

Выдай только готовый скрипт с маркерами [PAUSE] и [BEAT]. Без пояснений."""


def p_voiceover_style(sit, mode):
    """Направление голоса для ElevenLabs."""
    mode_id  = str(mode.get("ИД","Humor"))
    style    = VO_STYLES.get(mode_id,"Тёплый, разговорный.")
    duration = int(mode.get("Длит_сек",45) or 45)
    pace     = str(mode.get("Темп_ВО","medium") or "medium")

    return f"""ELEVENLABS — настройки голоса:

Язык: тайский, натуральный акцент
Стиль: {style}
Темп: {pace} — чёткая артикуляция, без спешки
Пол: на усмотрение (выбери что лучше подходит по тону скрипта)
Длительность: ~{duration} сек

СУБТИТРЫ: включить экспорт ✓ → скачать .srt → импортировать в CapCut

После генерации: прослушай один раз. Если тон не попал — перегенерируй перед скачиванием."""


def p_video_visual(sit, mode):
    """Промт для Kling / Runway — только визуал, без диалогов."""
    sit_text = sit.get("Ситуация","")
    industry = sit.get("Индустрия","")
    mode_id  = str(mode.get("ИД","Humor"))
    duration = int(mode.get("Длит_сек",45) or 45)
    pace     = str(mode.get("Темп_Видео","medium") or "medium")

    # Industry-specific visual cues
    env_map = {
        "restaurant":"тайское кафе или ресторан, яркий свет на кухне или кассе",
        "hotel":"стойка регистрации отеля, лобби, номер для гостей",
        "clinic":"регистратура клиники, кабинет врача, аптека",
        "ecommerce":"небольшой склад или домашний офис с коробками и упаковочным столом",
        "retail":"торговый зал магазина, стеллажи с товарами",
        "agency":"открытый офис, большие мониторы, переговорная",
        "freelancer":"домашний рабочий стол, ноутбук, кофе",
        "logistics":"склад или диспетчерская, карты доставок на экране",
        "education":"классная комната или учительский стол",
        "spa":"ресепшн спа или кабинет процедур",
        "fitness":"фитнес-зал, стойка регистрации",
        "manufacturing":"производственный цех, станки, рабочие места",
        "general":"офис малого бизнеса в Таиланде, современный но реальный",
    }
    env = env_map.get(industry,"офис или рабочее пространство малого бизнеса в ЮВА")

    mood_map = {
        "Humor":"слегка комедийный, преувеличенно напряжённый",
        "Serious":"серьёзный, задумчивый",
        "Demo":"динамичный, вдохновляющий",
        "Educational":"спокойный, поясняющий",
        "Before_After":"сначала хаотичный → в конце упорядоченный",
        "Myth_Busting":"уверенный",
        "Story":"кинематографический, с эмоцией",
        "POV":"от первого лица, субъективная камера",
    }
    mood = mood_map.get(mode_id,"профессиональный, живой")

    return f"""Vertical video 9:16, {duration} seconds. {pace} pacing.
Tool: Kling (primary) / Runway (fallback)

SCENE: {sit_text}
ENVIRONMENT: {env}
MOOD: {mood}

SHOT SEQUENCE:
[0-3s]   HOOK — peak of the problem at its most absurd or relatable moment
          Close-up: hands, screen, face showing the stress or confusion
[3-{duration-10}s] ESCALATION — the manual process in action
          Shots: typing, scrolling phone, stacks of paper, exhausted expression
          Clock or time cue showing late hour
[{duration-9}-{duration}s] RESOLUTION HINT — brief visual calm or contrast

VISUAL RULES:
• Southeast Asian business setting — realistic, not stock photo
• Real people, real expressions — no models or staged look
• Bright natural lighting in business environment
• NO text overlays (subtitles are separate)
• NO robots, AI chips, data holograms, Western-only offices

NOTE: Visuals only. All narration is handled separately via voiceover."""


def p_image(sit, mode, aspect, has_thai):
    """Промт для генератора изображений."""
    sit_text = sit.get("Ситуация","")
    industry = sit.get("Индустрия","")
    mode_id  = str(mode.get("ИД","Humor"))

    style_map = {
        "Humor":       "slightly exaggerated, comic energy, warm colors",
        "Serious":     "clean, documentary realism, muted tones",
        "Demo":        "dynamic, before-after split or transformation visual",
        "Educational": "clean infographic-style or friendly illustration",
        "Before_After":"split composition: chaotic left / ordered right",
        "Myth_Busting":"bold, confident, clear contrast",
        "Story":       "cinematic, emotional, character-driven",
        "POV":         "first-person perspective, personal, intimate",
    }
    style = style_map.get(mode_id,"clean modern illustration")
    thai_note = "\n⚠ THAI TEXT IN IMAGE: Yes. Use Ideogram v3 for this reason specifically." if has_thai else ""

    return f"""Modern {style} image for social media.
Aspect ratio: {aspect}
{thai_note}
SCENE: {sit_text}
INDUSTRY: {industry}

STYLE REQUIREMENTS:
• Southeast Asian business context — Thai or Lao environment
• Real human element: person experiencing the problem (not posing)
• Brand aesthetic: transformation from chaos to order
• Clean, contemporary — not stock photo look

INCLUDE: The core visual tension from the situation — the pile of papers,
the overwhelming screen, the exhausted face, the chaos

AVOID: Robots, AI brain/chip imagery, data holograms, generic Western
corporate settings, anything that looks like AI marketing material"""


def p_caption(sit, mode, platform_cfg, char_name=""):
    """Промт для подписи — идёт в Claude или GPT."""
    platform     = platform_cfg.get("Платформа","")
    lang         = str(platform_cfg.get("Язык",""))
    max_chars    = platform_cfg.get("Символы_макс","")
    sit_text     = sit.get("Ситуация","")
    hook         = sit.get("Хук","")
    industry     = sit.get("Индустрия","")
    mode_id      = str(mode.get("ИД","Humor"))
    mode_name    = str(mode.get("Режим","Юмор"))
    mode_inst    = MODE_INSTRUCTIONS.get(mode_id,"")
    arc          = EMOTION_ARCS.get(mode_id,"нарастание → вывод")

    # Language instruction
    if "Тайский" in lang and "English" in lang:
        lang_instr = ("Пиши на тайском (основной язык, разговорный стиль).\n"
                      "После текста добавь разделитель ▼ и сокращённый перевод на английский (тот же тон, не дословно).")
    elif "Тайский" in lang:
        lang_instr = "Пиши только на тайском. Разговорный стиль, не официальный."
    else:
        lang_instr = "Write in English. Conversational, direct — not corporate."

    hook_line = f"\nХУК (отправная точка для первой строки, можно развить):\n{hook}" if hook else ""
    char_note = f"\nПЕРСОНАЖ: {char_name} — пиши от его лица или с его точки зрения.\n" if char_name else ""
    limit     = f"\nЛИМИТ: {max_chars} символов." if max_chars else ""

    # Platform-specific structural note
    struct_map = {
        "TikTok":    "Структура: хука (1-2 строки) → суть (2-4 строки) → финал. Коротко и хлёстко.",
        "Facebook":  "Структура: хука → развитие (4-7 строк) → финал с вопросом. Можно длиннее.",
        "Instagram": "Структура: хука (стоп-скролл) → тело (4-6 строк) → финал → ▼ EN → 3-5 хэштегов.",
        "Telegram":  "Структура: хука → развёрнутый анализ (6-10 строк) → чёткий вывод. Тон серьёзнее.",
        "YouTube":   "Структура: заголовок (до 70 символов) → описание (150-300 символов). Добавь 3-5 ключевых слова.",
        "Line":      "Структура: хука → суть (3-5 строк) → финал. Очень коротко, только главное.",
    }
    struct = struct_map.get(platform,"Хука → развитие → финал.")

    return f"""Напиши подпись к посту для AI Flow Studio.

БРЕНД:
{BRAND_BLOCK}

СИТУАЦИЯ ДЛЯ КОНТЕНТА:
{sit_text}
{hook_line}
{char_note}
РЕЖИМ: {mode_name}
{mode_inst}

ПЛАТФОРМА: {platform}
{lang_instr}

{struct}
Эмоциональная дуга: {arc}
{limit}

ГОЛОС БРЕНДА — НЕЛЬЗЯ:
- «ИИ изменит ваш бизнес» / «инновационные AI-решения» / «синергия»
- Унижать людей (смеёмся над абсурдом ситуации, не над человеком)
- Жёсткие продающие призывы в конце
- Корпоративный язык и пустые слова

ГОЛОС БРЕНДА — НУЖНО:
- Узнаваемые конкретные детали (название инструментов, числа, время)
- Финал: открытый вопрос, мягкий инсайт или «а у вас так же?»
- Разговорно, как будто пишет живой человек — не бот

Выдай только готовый текст подписи. Без пояснений и ремарок."""


# ── Excel output ──────────────────────────────────────────────────
def write_day_hdr(ws, row, date_str, day_num, hex_c):
    ws.row_dimensions[row].height = 22
    c = ws.cell(row=row, column=1, value=f"  ДЕНЬ {day_num}   —   {date_str}")
    c.fill = mkfill(hex_c); c.font = mkfont(True,12,"FFFFFF")
    c.alignment = Alignment(vertical="center",horizontal="left")
    for col in range(2, len(COLS)+1):
        ws.cell(row=row,column=col).fill = mkfill(hex_c)

def write_atom_hdr(ws, row, date_str, n, mode_name, industry, situation, hex_c):
    ws.row_dimensions[row].height = 26
    vals=[date_str,f"#{n}",f"[{mode_name}] {industry}","","◀ СИТУАЦИЯ",situation,"",""]
    for i,v in enumerate(vals,1):
        c=ws.cell(row=row,column=i,value=v)
        c.fill=mkfill(hex_c); c.font=mkfont(True,10)
        c.alignment=mkalign("center" if i in (1,2,4,7,8) else "left")

def write_row(ws, row, date_str, n, step, platform, task, prompt, tool,
              prod_hex, task_hex, height=100):
    ws.row_dimensions[row].height = height
    data=[date_str,f"#{n}",step,platform,task,prompt,tool,""]
    fills=[mkfill(prod_hex)]*5+[mkfill(task_hex or prod_hex),mkfill(prod_hex),mkfill(WHITE)]
    for i,(v,f) in enumerate(zip(data,fills),1):
        c=ws.cell(row=row,column=i,value=v)
        c.fill=f; c.font=mkfont(size=9 if i==6 else 10)
        c.alignment=mkalign("center" if i in (1,2,4,7,8) else "left")

def build_excel(output_data):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Промты"
    for i,(header,width) in enumerate(COLS,1):
        ws.column_dimensions[get_column_letter(i)].width=width
        c=ws.cell(row=1,column=i,value=header)
        c.fill=mkfill("1E293B"); c.font=mkfont(True,10,"FFFFFF"); c.alignment=mkcenter()
    ws.row_dimensions[1].height=22; ws.freeze_panes="A2"

    row=2
    for day_idx, day in enumerate(output_data):
        pal=PALETTE[day_idx%3]
        day_h,atom_bg,prod_bg=pal
        write_day_hdr(ws,row,day["date"],day_idx+1,day_h); row+=1

        for atom_idx, atom in enumerate(day["atoms"]):
            n=atom_idx+1
            narration = atom["narration"]
            needs_vo  = narration in ("narrated","both")
            has_video = "video" in atom["format_fit"]

            sit_short=(atom["situation"] or "")[:90]
            write_atom_hdr(ws,row,day["date"],n,atom["mode_name"],atom["industry"],sit_short,atom_bg)
            row+=1

            # 1. Текст озвучки
            vo_prompt = atom["vo_script"] if needs_vo else "— не нужно: карусель без озвучки"
            write_row(ws,row,day["date"],n,"Производство","ВСЕ ПЛАТФОРМЫ","Текст озвучки",
                      vo_prompt,atom["text_tool"],prod_bg,TASK_CLR["Текст озвучки"],130)
            row+=1

            # 2. Стиль озвучки
            vs_prompt = atom["vo_style"] if needs_vo else "— не нужно"
            write_row(ws,row,day["date"],n,"Производство","ElevenLabs","Стиль озвучки",
                      vs_prompt,atom["vo_tool"],prod_bg,TASK_CLR["Стиль озвучки"],90)
            row+=1

            # 3. Промт видео
            vid_prompt = atom["video_prompt"] if has_video else "— не нужно для этого формата"
            write_row(ws,row,day["date"],n,"Производство","TikTok/IG/FB/YT","Промт видео",
                      vid_prompt,atom["video_tool"],prod_bg,TASK_CLR["Промт видео"],130)
            row+=1

            # 4. Промт картинки
            write_row(ws,row,day["date"],n,"Производство","Все платформы","Промт картинки",
                      atom["image_prompt"],atom["image_tool"],prod_bg,TASK_CLR["Промт картинки"],100)
            row+=1

            # 5-10. Подписи по платформам
            for p in atom["captions"]:
                write_row(ws,row,day["date"],n,"Подпись",p["platform"],"Подпись к посту",
                          p["prompt"],p["text_tool"],prod_bg,TASK_CLR["Подпись"],130)
                row+=1

            ws.row_dimensions[row].height=6; row+=1

        ws.row_dimensions[row].height=10; row+=1
    return wb

# ── log update ────────────────────────────────────────────────────
def update_log(entries):
    try:
        wb=openpyxl.load_workbook(DNA_FILE)
        if "Журнал" not in wb.sheetnames:
            ws=wb.create_sheet("Журнал")
            ws.append(["Дата","ИД_Ситуации","Режим","Индустрия","Персонаж"])
        else:
            ws=wb["Журнал"]
        for e in entries:
            ws.append([e["date"],e["sit_id"],e["mode"],e["industry"],e.get("char","")])
        wb.save(DNA_FILE); wb.close(); return True
    except PermissionError:
        print(f"  ВНИМАНИЕ: {DNA_FILE} открыт в Excel — журнал не обновлён"); return False
    except Exception as ex:
        print(f"  ВНИМАНИЕ: ошибка журнала: {ex}"); return False

# ── main ──────────────────────────────────────────────────────────
def main():
    horizon=int(sys.argv[1]) if len(sys.argv)>1 else DEFAULT_DAYS
    print(f"\n{'='*56}\n  AI Flow Studio — Генератор промтов v2.0\n{'='*56}")

    if not Path(DNA_FILE).exists():
        print(f"\n  ОШИБКА: {DNA_FILE} не найден. Сначала запусти setup_dna_ru.py\n"); sys.exit(1)

    print(f"\n  Загружаю {DNA_FILE}...")
    wb=openpyxl.load_workbook(DNA_FILE,read_only=True)
    situations  = load_sheet(wb,"Ситуации")
    modes_raw   = load_sheet(wb,"Режимы")
    platforms   = load_sheet(wb,"Платформы")
    tools       = load_sheet(wb,"Инструменты")
    log_raw     = load_sheet(wb,"Журнал")
    wb.close()

    if not situations:
        print("  ОШИБКА: лист Ситуации пуст.\n"); sys.exit(1)

    modes_idx = {str(m.get("ИД","")): m for m in modes_raw}
    tmap      = tool_map(tools)
    history   = parse_log(log_raw)

    print(f"  Ситуации: {len(situations)}  Режимы: {len(modes_raw)}  "
          f"Платформы: {len(platforms)}  Журнал: {len(history)}")
    print(f"\n  Генерирую {horizon} дн. × {ATOMS_PER_DAY} атома...\n")

    start     = datetime.now().replace(hour=0,minute=0,second=0,microsecond=0)
    out_data  = []
    new_log   = []

    for day_off in range(horizon):
        target   = start+timedelta(days=day_off)
        date_str = target.strftime("%Y-%m-%d")
        combined = history+[{"date":target,"sit_id":e["sit_id"],"mode":e["mode"],"industry":e["industry"]} for e in new_log]
        atoms_raw= select_atoms(situations,modes_idx,combined,target)
        print(f"  День {day_off+1}  ({date_str})")

        day_atoms=[]
        for idx,sit in enumerate(atoms_raw):
            mode_id   = str(sit.get("Режим","Humor"))
            mode      = modes_idx.get(mode_id, modes_raw[0] if modes_raw else {})
            sit_id    = str(sit.get("ИД",""))
            industry  = str(sit.get("Индустрия","general"))
            char_id   = str(sit.get("Персонаж","") or "")
            mode_name = str(mode.get("Режим",mode_id))
            narration = str(sit.get("Нарация","narrated"))
            fmt_fit   = str(sit.get("Форматы","video")).lower()
            has_thai  = True  # платформы TH+EN доминируют

            print(f"    #{idx+1} [{mode_id:13s}] {industry:14s} "
                  f"{(sit.get('Ситуация',''))[:46]}...")

            # core production prompts
            vo_script    = p_voiceover_script(sit, mode)
            vo_style     = p_voiceover_style(sit, mode)
            video_prompt = p_video_visual(sit, mode)
            img_prompt   = p_image(sit, mode,
                                   (platforms[0].get("Соотношение","9:16") if platforms else "9:16"),
                                   has_thai)
            img_tool     = pick_image_tool(tools, has_thai)

            # captions per platform
            captions=[]
            for p_cfg in platforms:
                captions.append({
                    "platform":  p_cfg.get("Платформа",""),
                    "prompt":    p_caption(sit, mode, p_cfg, char_id),
                    "text_tool": tmap.get("text","Claude"),
                })

            day_atoms.append({
                "sit_id":       sit_id,
                "situation":    sit.get("Ситуация",""),
                "mode_id":      mode_id,
                "mode_name":    mode_name,
                "industry":     industry,
                "narration":    narration,
                "format_fit":   fmt_fit,
                "vo_script":    vo_script,
                "vo_style":     vo_style,
                "vo_tool":      tmap.get("voiceover","ElevenLabs"),
                "video_prompt": video_prompt,
                "video_tool":   tmap.get("video","Kling"),
                "image_prompt": img_prompt,
                "image_tool":   img_tool,
                "text_tool":    tmap.get("text","Claude"),
                "captions":     captions,
            })
            new_log.append({"date":date_str,"sit_id":sit_id,"mode":mode_id,"industry":industry,"char":char_id})

        out_data.append({"date":date_str,"atoms":day_atoms})

    # write output
    OUTPUT_DIR.mkdir(exist_ok=True)
    out_path=OUTPUT_DIR/f"prompts_{start.strftime('%Y%m%d')}.xlsx"
    print(f"\n  Записываю Excel...")
    out_wb=build_excel(out_data)
    try:
        out_wb.save(out_path)
    except PermissionError:
        alt=OUTPUT_DIR/f"prompts_{start.strftime('%Y%m%d_%H%M%S')}.xlsx"
        out_wb.save(alt); out_path=alt

    update_log(new_log)

    total=horizon*ATOMS_PER_DAY
    print(f"\n  ✓  {out_path}")
    print(f"\n  Итог: {horizon} дн. × {ATOMS_PER_DAY} атома = {total} контент-единицы")
    print(f"  На атом: озвучка + видео + картинка + {len(platforms)} подписей")
    print(f"\n  Порядок работы: Текст озвучки → ElevenLabs → Kling → картинка → подписи")
    print(f"{'='*56}\n")

if __name__=="__main__":
    main()
